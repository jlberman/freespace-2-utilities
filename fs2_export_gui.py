#!/usr/bin/env python3
"""
fs2_export_gui.py
GUI wrapper around the FS2 message metadata extractor.
Supports individual file selection and folder batch mode.

Requirements: Python 3.8+, openpyxl  (pip install openpyxl)
              tkinter (included with standard Python on Windows/macOS/most Linux)

Usage:  python3 fs2_export_gui.py
"""

import os
import re
import sys
import threading
from pathlib import Path

# ── optional openpyxl check before launching GUI ─────────────────────────────
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# ═════════════════════════════════════════════════════════════════════════════
#  EXTRACTION ENGINE  (embedded from fs2_message_export.py)
# ═════════════════════════════════════════════════════════════════════════════

def _extract_messages_section(content):
    start = content.find('#Messages')
    if start == -1:
        return None
    nxt = re.search(r'\n#(?!Messages)\w', content[start + 1:])
    end = start + 1 + nxt.start() if nxt else len(content)
    return content[start:end]


def _parse_senders(content):
    senders = {}

    def add(msg, sender):
        senders.setdefault(msg, set()).add(sender)

    for sender, pri, msg in re.findall(
        r'\(\s*send-message\s+"([^"]+)"\s+"([^"]+)"\s+"([^"]+)"\s*\)', content):
        add(msg, sender)

    for m in re.finditer(r'\(\s*send-message-list\s+([\s\S]*?)\n\s*\)', content):
        tokens = re.findall(r'"([^"]*)"|\b(\d+)\b', m.group(1))
        flat = [t[0] if t[0] != '' else t[1] for t in tokens]
        i = 0
        while i + 3 <= len(flat):
            try:
                int(flat[i + 3]); add(flat[i + 2], flat[i]); i += 4
            except (ValueError, IndexError):
                i += 1

    for m in re.finditer(r'\(\s*send-message-chain\s+([\s\S]*?)\n\s*\)', content):
        tokens = re.findall(r'"([^"]*)"|\b(\d+)\b', m.group(1))
        flat = [t[0] if t[0] != '' else t[1] for t in tokens]
        i = 1
        while i + 3 <= len(flat):
            try:
                int(flat[i + 3]); add(flat[i + 2], flat[i]); i += 4
            except (ValueError, IndexError):
                i += 1

    return senders


def _relabel_sender(sender, avi_name):
    if sender == '#Command':
        avi = avi_name.lower()
        if avi == 'head-fr04':
            return 'Karpinsky'
        if avi == 'head-merc01':
            return 'Tejada'
    return sender


def _mission_sort_key(num_str):
    m = re.match(r'(\d+)(\w*)', str(num_str))
    return (int(m.group(1)), m.group(2).lower()) if m else (9999, str(num_str))


def extract_file(fpath):
    """Parse one .fs2 file. Returns (mission_num, mission_name, rows)."""
    try:
        content = Path(fpath).read_text(encoding='utf-8', errors='replace')
    except OSError as e:
        return None, None, [], str(e)

    content = content.replace('\r\n', '\n').replace('\r', '\n')

    mn_match = re.search(r'\$Name:\s+XSTR\("([^"]+)"', content)
    mission_name = mn_match.group(1) if mn_match else Path(fpath).stem

    num_match = re.search(r'INFRC3-(\d+\w*)', fpath, re.IGNORECASE)
    if not num_match:
        num_match = re.search(r'(\d+\w*)\.fs2$', fpath, re.IGNORECASE)
    mission_num = num_match.group(1) if num_match else Path(fpath).stem

    section = _extract_messages_section(content)
    if section is None:
        return mission_num, mission_name, [], None

    senders = _parse_senders(content)

    cmd_persona_match = re.search(r'\$Command Persona:\s*(.+)', section)
    cmd_persona = cmd_persona_match.group(1).strip() if cmd_persona_match else ''

    rows = []
    for block in section.split('\n$Name:')[1:]:
        blines = block.strip().splitlines()
        if not blines:
            continue
        msg_name = blines[0].strip()

        text_match = re.search(r'\$MessageNew:\s+XSTR\("([^"]+)"', block)
        msg_text = text_match.group(1).strip() if text_match else ''

        avi_match = re.search(r'\+AVI Name:\s*(.+)', block)
        avi_name = avi_match.group(1).strip() if avi_match else ''

        wave_match = re.search(r'\+Wave Name:\s*(.+)', block)
        wave_name = wave_match.group(1).strip() if wave_match else ''

        persona_match = re.search(r'\+Persona:\s*(.+)', block)
        persona = persona_match.group(1).strip() if persona_match else ''

        sender_set = senders.get(msg_name, set())
        sender_str = ', '.join(sorted(
            _relabel_sender(s, avi_name) for s in sender_set
        )) if sender_set else ''

        rows.append({
            'mission_num':  mission_num,
            'mission_name': mission_name,
            'msg_name':     msg_name,
            'sender':       sender_str,
            'avi_name':     avi_name,
            'wave_name':    wave_name,
            'persona':      persona,
            'cmd_persona':  cmd_persona,
            'msg_text':     msg_text,
        })

    return mission_num, mission_name, rows, None


# ── Spreadsheet builder ───────────────────────────────────────────────────────

HEADERS = [
    'Mission #', 'Mission Name', 'Message Name ($Name)',
    'Sender', 'AVI Name', 'Wave Name',
    'Persona (+Persona)', 'Command Persona', 'Message Text',
]

_C_HDR_BG  = '1A2744'
_C_HDR_FG  = 'E8F0FE'
_C_MIS_BG  = '0D3349'
_C_ALT     = 'F2F6FA'
_C_WHITE   = 'FFFFFF'
_C_ACCENT  = '00C89C'
_C_WARN    = 'FFF3CD'


def _thin():
    s = Side(style='thin', color='C8D0DC')
    return Border(left=s, right=s, top=s, bottom=s)


def _write_sheet(ws, rows, start_row=1):
    hr = start_row
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(hr, ci, h)
        c.font = Font(name='Arial', bold=True, size=9, color=_C_HDR_FG)
        c.fill = PatternFill('solid', fgColor=_C_ACCENT if h == 'Sender' else _C_HDR_BG)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _thin()
    ws.row_dimensions[hr].height = 28
    ws.freeze_panes = ws.cell(hr + 1, 1)
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f'A{hr}:{last_col}{hr + len(rows)}'

    prev = None
    for i, r in enumerate(rows):
        ri = hr + 1 + i
        mc = r['mission_num'] != prev
        prev = r['mission_num']
        vals = [r['mission_num'], r['mission_name'], r['msg_name'], r['sender'],
                r['avi_name'], r['wave_name'], r['persona'], r['cmd_persona'], r['msg_text']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name='Arial', size=9,
                          bold=(ci == 4 and bool(v)))
            c.border = _thin()
            c.alignment = Alignment(vertical='top', wrap_text=(ci == 9))
            if not r['sender']:
                c.fill = PatternFill('solid', fgColor=_C_WARN)
            else:
                c.fill = PatternFill('solid', fgColor=_C_ALT if i % 2 == 0 else _C_WHITE)
        ws.row_dimensions[ri].height = 15

    for i, w in enumerate([10,26,32,18,14,32,14,24,70], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(all_rows):
    from collections import defaultdict
    wb = Workbook()

    # Summary
    ws_s = wb.active
    ws_s.title = 'Summary'
    ws_s['A1'] = 'FS2 Message Export'
    ws_s['A1'].font = Font(name='Arial', bold=True, size=14, color=_C_HDR_BG)
    ws_s['A2'] = f'Total messages: {len(all_rows)}'
    ws_s['A2'].font = Font(name='Arial', size=10, color='555555')
    ws_s['A3'] = f'With confirmed sender: {sum(1 for r in all_rows if r["sender"])}'
    ws_s['A3'].font = Font(name='Arial', size=10, color='555555')

    for ci, h in enumerate(['Mission', 'Total', 'With Sender', 'Without Sender'], 1):
        c = ws_s.cell(5, ci, h)
        c.font = Font(name='Arial', bold=True, color=_C_HDR_FG)
        c.fill = PatternFill('solid', fgColor=_C_HDR_BG)
        c.alignment = Alignment(horizontal='center')

    by_m = defaultdict(lambda: {'name': '', 'total': 0, 'with': 0})
    for r in all_rows:
        k = r['mission_num']
        by_m[k]['name'] = r['mission_name']
        by_m[k]['total'] += 1
        if r['sender']:
            by_m[k]['with'] += 1

    sorted_m = sorted(by_m.items(), key=lambda x: _mission_sort_key(x[0]))
    for i, (mn, d) in enumerate(sorted_m, 6):
        for ci, v in enumerate([f"M{mn}: {d['name']}", d['total'], d['with'],
                                 d['total'] - d['with']], 1):
            c = ws_s.cell(i, ci, v)
            c.font = Font(name='Arial', size=9)
            c.border = _thin()
            c.fill = PatternFill('solid', fgColor=_C_ALT if i % 2 == 0 else _C_WHITE)
            if ci > 1:
                c.alignment = Alignment(horizontal='center')

    tr = len(sorted_m) + 6
    for ci, v in enumerate(['TOTAL', f'=SUM(B6:B{tr-1})',
                             f'=SUM(C6:C{tr-1})', f'=SUM(D6:D{tr-1})'], 1):
        c = ws_s.cell(tr, ci, v)
        c.font = Font(name='Arial', bold=True, color=_C_HDR_FG)
        c.fill = PatternFill('solid', fgColor=_C_MIS_BG)
        c.border = _thin()
        if ci > 1:
            c.alignment = Alignment(horizontal='center')

    for col, w in zip('ABCD', [38, 18, 18, 18]):
        ws_s.column_dimensions[col].width = w

    # All messages
    ws_all = wb.create_sheet('All Messages')
    _write_sheet(ws_all, all_rows)

    # Per-mission
    for mn, d in sorted_m:
        m_rows = [r for r in all_rows if r['mission_num'] == mn]
        sname = re.sub(r'[\\/*?:\[\]]', '', f'M{mn}')[:31]
        ws_m = wb.create_sheet(sname)
        ws_m.merge_cells('A1:I1')
        ws_m['A1'] = f"Mission {mn}: {d['name']}"
        ws_m['A1'].font = Font(name='Arial', bold=True, size=11, color=_C_HDR_FG)
        ws_m['A1'].fill = PatternFill('solid', fgColor=_C_MIS_BG)
        ws_m['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws_m.row_dimensions[1].height = 22
        _write_sheet(ws_m, m_rows, start_row=2)

    return wb


# ═════════════════════════════════════════════════════════════════════════════
#  DESIGN TOKENS
# ═════════════════════════════════════════════════════════════════════════════

BG          = '#0e1118'
SURFACE     = '#161b27'
SURFACE2    = '#1e2535'
BORDER_C    = '#2a3347'
ACCENT      = '#3dd68c'
ACCENT2     = '#1a5c3d'
TEXT        = '#ccd6e8'
TEXT_DIM    = '#4e5f78'
TEXT_BRIGHT = '#eef3ff'
DANGER      = '#e05252'
WARN_C      = '#e09820'
SUCCESS     = '#3dd68c'

IS_WIN  = sys.platform == 'win32'
IS_MAC  = sys.platform == 'darwin'
F_BODY  = ('Segoe UI', 10) if IS_WIN else ('SF Pro Text', 10) if IS_MAC else ('Helvetica', 10)
F_MONO  = ('Consolas', 9)  if IS_WIN else ('Menlo', 9)       if IS_MAC else ('Courier', 9)
F_TITLE = (F_BODY[0], 13, 'bold')
F_SMALL = (F_BODY[0], 8)
F_CAP   = (F_BODY[0], 8, 'bold')


# ═════════════════════════════════════════════════════════════════════════════
#  WIDGETS
# ═════════════════════════════════════════════════════════════════════════════

class FlatButton(tk.Button):
    def __init__(self, parent, text, command, accent=False, danger=False,
                 small=False, icon='', **kw):
        bg = ACCENT if accent else (DANGER if danger else SURFACE2)
        fg = '#0e1118' if accent else TEXT
        hov = '#52eaaa' if accent else ('#e87070' if danger else BORDER_C)
        font = F_SMALL if small else F_BODY
        lbl = f'{icon} {text}'.strip() if icon else text
        super().__init__(parent, text=lbl, command=command,
                         bg=bg, fg=fg, activebackground=hov, activeforeground=fg,
                         relief='flat', bd=0, padx=10, pady=5,
                         font=font, cursor='hand2', **kw)
        self._bg, self._hov = bg, hov
        self.bind('<Enter>', lambda e: self.config(bg=self._hov))
        self.bind('<Leave>', lambda e: self.config(bg=self._bg))

    def set_accent(self, on):
        self._bg = ACCENT if on else SURFACE2
        self._hov = '#52eaaa' if on else BORDER_C
        self.config(bg=self._bg, fg='#0e1118' if on else TEXT)


class SectionLabel(tk.Frame):
    def __init__(self, parent, text):
        super().__init__(parent, bg=BG)
        tk.Label(self, text=text.upper(), font=F_CAP,
                 fg=ACCENT, bg=BG).pack(side='left', padx=(0, 8))
        tk.Frame(self, bg=ACCENT2, height=1).pack(side='left', fill='x', expand=True)


class FileRow(tk.Frame):
    """Single removable file row in the file list."""

    def __init__(self, parent, fpath, on_remove, even=True):
        bg = SURFACE if even else SURFACE2
        super().__init__(parent, bg=bg, pady=2)
        self.fpath = fpath
        name = Path(fpath).name
        size_kb = _safe_size(fpath)

        tk.Label(self, text='📄', font=F_BODY, bg=bg, fg=TEXT_DIM,
                 width=2).pack(side='left', padx=(6, 2))
        tk.Label(self, text=name, font=F_MONO, bg=bg, fg=TEXT,
                 anchor='w').pack(side='left', fill='x', expand=True)
        tk.Label(self, text=size_kb, font=F_SMALL, bg=bg,
                 fg=TEXT_DIM, width=8).pack(side='left', padx=4)
        FlatButton(self, '×', lambda: on_remove(fpath),
                   small=True, icon='').pack(side='right', padx=4)


def _safe_size(fpath):
    try:
        kb = os.path.getsize(fpath) / 1024
        return f'{kb:.0f} KB'
    except Exception:
        return ''


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═════════════════════════════════════════════════════════════════════════════

class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title('FS2 Message Exporter')
        self.configure(bg=BG)
        self.minsize(820, 560)
        self.geometry('1020x680')

        self._files = []          # list of absolute path strings
        self._result_rows = []    # extracted data rows
        self._running = False

        self.out_path = tk.StringVar()
        self.var_warnings = tk.BooleanVar(value=True)

        self._build()
        self._check_openpyxl()

    # ── openpyxl gate ──────────────────────────────────────────────────────

    def _check_openpyxl(self):
        if not HAVE_OPENPYXL:
            self._log('ERROR: openpyxl is not installed.', 'err')
            self._log('Run:  pip install openpyxl', 'dim')
            self.run_btn.config(state='disabled')

    # ── Build UI ───────────────────────────────────────────────────────────

    def _build(self):
        # ── Top bar ───────────────────────────────────────────────────────
        topbar = tk.Frame(self, bg=SURFACE, height=50)
        topbar.pack(fill='x')
        topbar.pack_propagate(False)
        tk.Label(topbar, text='FS2  MESSAGE  EXPORTER',
                 font=F_TITLE, fg=ACCENT, bg=SURFACE).pack(side='left', padx=18, pady=14)
        tk.Label(topbar, text='·  xlsx metadata extractor',
                 font=F_SMALL, fg=TEXT_DIM, bg=SURFACE).pack(side='left', pady=18)

        # ── Body (left panel + right log) ─────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill='both', expand=True)

        left = tk.Frame(body, bg=BG, width=420)
        left.pack(side='left', fill='y', padx=(12, 6), pady=10)
        left.pack_propagate(False)

        right = tk.Frame(body, bg=BG)
        right.pack(side='left', fill='both', expand=True, padx=(0, 12), pady=10)

        self._build_left(left)
        self._build_right(right)

        # ── Status bar ────────────────────────────────────────────────────
        sbar = tk.Frame(self, bg=SURFACE, height=26)
        sbar.pack(fill='x', side='bottom')
        sbar.pack_propagate(False)
        self.status_var = tk.StringVar(value='Ready.')
        tk.Label(sbar, textvariable=self.status_var,
                 font=F_SMALL, fg=TEXT_DIM, bg=SURFACE,
                 anchor='w').pack(side='left', padx=10, pady=4)
        self.count_var = tk.StringVar(value='0 files')
        tk.Label(sbar, textvariable=self.count_var,
                 font=F_SMALL, fg=TEXT_DIM, bg=SURFACE,
                 anchor='e').pack(side='right', padx=10, pady=4)

    # ── Left panel ─────────────────────────────────────────────────────────

    def _build_left(self, parent):

        # ── File input mode tabs ──────────────────────────────────────────
        SectionLabel(parent, 'Add Files').pack(fill='x', pady=(0, 4))

        tab_frame = tk.Frame(parent, bg=BG)
        tab_frame.pack(fill='x')

        self._mode = tk.StringVar(value='files')
        btn_files = FlatButton(tab_frame, 'Pick Files', self._add_files)
        btn_files.pack(side='left', padx=(0, 2))
        btn_folder = FlatButton(tab_frame, 'Pick Folder', self._add_folder)
        btn_folder.pack(side='left', padx=(0, 2))
        FlatButton(tab_frame, 'Clear All', self._clear_files,
                   danger=True).pack(side='right')

        # ── File list ─────────────────────────────────────────────────────
        list_outer = tk.Frame(parent, bg=SURFACE,
                              highlightthickness=1,
                              highlightbackground=BORDER_C)
        list_outer.pack(fill='both', expand=True, pady=(6, 0))

        self._canvas = tk.Canvas(list_outer, bg=SURFACE,
                                 highlightthickness=0, bd=0)
        self._vsb = tk.Scrollbar(list_outer, orient='vertical',
                                 command=self._canvas.yview,
                                 bg=SURFACE, troughcolor=SURFACE,
                                 relief='flat', bd=0)
        self._canvas.configure(yscrollcommand=self._vsb.set)
        self._vsb.pack(side='right', fill='y')
        self._canvas.pack(side='left', fill='both', expand=True)

        self._list_frame = tk.Frame(self._canvas, bg=SURFACE)
        self._canvas_window = self._canvas.create_window(
            (0, 0), window=self._list_frame, anchor='nw')

        self._list_frame.bind('<Configure>', self._on_list_resize)
        self._canvas.bind('<Configure>', self._on_canvas_resize)
        self._canvas.bind_all('<MouseWheel>', self._on_scroll)

        # empty state label
        self._empty_lbl = tk.Label(self._list_frame,
                                   text='Drop files here or use buttons above',
                                   font=F_SMALL, fg=TEXT_DIM, bg=SURFACE,
                                   pady=20)
        self._empty_lbl.pack()

        # ── Output path ───────────────────────────────────────────────────
        SectionLabel(parent, 'Output').pack(fill='x', pady=(10, 4))

        out_row = tk.Frame(parent, bg=BG)
        out_row.pack(fill='x')

        self.out_entry = tk.Entry(out_row, textvariable=self.out_path,
                                  bg=SURFACE2, fg=TEXT,
                                  insertbackground=ACCENT,
                                  relief='flat', bd=6, font=F_MONO)
        self.out_entry.pack(side='left', fill='x', expand=True)
        FlatButton(out_row, '…', self._pick_output,
                   small=True).pack(side='left', padx=(4, 0))

        # ── Options ───────────────────────────────────────────────────────
        SectionLabel(parent, 'Options').pack(fill='x', pady=(10, 4))

        opt_row = tk.Frame(parent, bg=BG)
        opt_row.pack(fill='x')
        cb = tk.Checkbutton(opt_row, text='Highlight rows without confirmed sender',
                            variable=self.var_warnings,
                            bg=BG, fg=TEXT, selectcolor=ACCENT2,
                            activebackground=BG, activeforeground=TEXT_BRIGHT,
                            font=F_BODY, relief='flat', bd=0)
        cb.pack(anchor='w')

        # ── Run ───────────────────────────────────────────────────────────
        tk.Frame(parent, bg=BORDER_C, height=1).pack(fill='x', pady=8)

        self.run_btn = FlatButton(parent, 'Export to Excel', self._run, accent=True)
        self.run_btn.pack(fill='x', ipady=6)

        self.progress = ttk.Progressbar(parent, mode='determinate', maximum=100)
        style = ttk.Style()
        style.theme_use('default')
        style.configure('TProgressbar', troughcolor=SURFACE2,
                        background=ACCENT, bordercolor=BORDER_C,
                        lightcolor=ACCENT, darkcolor=ACCENT)
        self.progress.pack(fill='x', pady=(5, 0))

    # ── Right panel (log + preview) ────────────────────────────────────────

    def _build_right(self, parent):
        SectionLabel(parent, 'Activity Log').pack(fill='x', pady=(0, 4))

        log_frame = tk.Frame(parent, bg=SURFACE,
                             highlightthickness=1,
                             highlightbackground=BORDER_C)
        log_frame.pack(fill='both', expand=True)

        xsb = tk.Scrollbar(log_frame, orient='horizontal',
                            bg=SURFACE, troughcolor=SURFACE, relief='flat', bd=0)
        xsb.pack(side='bottom', fill='x')
        ysb = tk.Scrollbar(log_frame, bg=SURFACE,
                           troughcolor=SURFACE, relief='flat', bd=0)
        ysb.pack(side='right', fill='y')

        self.log = tk.Text(log_frame, bg=SURFACE, fg=TEXT,
                           insertbackground=ACCENT, relief='flat', bd=8,
                           font=F_MONO, wrap='none', state='disabled',
                           xscrollcommand=xsb.set, yscrollcommand=ysb.set)
        self.log.pack(fill='both', expand=True)
        xsb.config(command=self.log.xview)
        ysb.config(command=self.log.yview)

        self.log.tag_configure('ok',  foreground=SUCCESS)
        self.log.tag_configure('err', foreground=DANGER)
        self.log.tag_configure('dim', foreground=TEXT_DIM)
        self.log.tag_configure('hdr', foreground=ACCENT,
                               font=(F_MONO[0], F_MONO[1], 'bold'))
        self.log.tag_configure('warn', foreground=WARN_C)

        # Result summary bar
        self.result_var = tk.StringVar(value='')
        tk.Label(parent, textvariable=self.result_var,
                 font=F_SMALL, fg=TEXT_DIM, bg=BG,
                 anchor='w').pack(fill='x', pady=(5, 0))

        # Save log button
        FlatButton(parent, 'Save Log…', self._save_log,
                   small=True).pack(anchor='w', pady=(4, 0))

    # ── File management ────────────────────────────────────────────────────

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title='Select .fs2 mission files',
            filetypes=[('FreeSpace 2 Mission', '*.fs2'), ('All files', '*.*')])
        self._add_paths(list(paths))

    def _add_folder(self):
        folder = filedialog.askdirectory(title='Select folder containing .fs2 files')
        if not folder:
            return
        found = sorted(str(p) for p in Path(folder).rglob('*.fs2'))
        if not found:
            messagebox.showinfo('No files found',
                                f'No .fs2 files found in:\n{folder}')
            return
        self._add_paths(found)

    def _add_paths(self, paths):
        existing = set(self._files)
        added = 0
        for p in paths:
            ap = str(Path(p).resolve())
            if ap not in existing:
                self._files.append(ap)
                existing.add(ap)
                added += 1
        if added:
            self._refresh_list()
            self._log(f'Added {added} file(s). Total: {len(self._files)}', 'ok')
        self._update_count()

    def _remove_file(self, fpath):
        if fpath in self._files:
            self._files.remove(fpath)
            self._refresh_list()
            self._update_count()
            self._log(f'Removed: {Path(fpath).name}', 'dim')

    def _clear_files(self):
        if self._files and messagebox.askyesno(
                'Clear file list', f'Remove all {len(self._files)} files?'):
            self._files.clear()
            self._refresh_list()
            self._update_count()
            self._log('File list cleared.', 'dim')

    def _refresh_list(self):
        for w in self._list_frame.winfo_children():
            w.destroy()
        if not self._files:
            lbl = tk.Label(self._list_frame,
                           text='No files selected — use buttons above to add .fs2 files',
                           font=F_SMALL, fg=TEXT_DIM, bg=SURFACE, pady=20)
            lbl.pack()
        else:
            for i, fpath in enumerate(self._files):
                row = FileRow(self._list_frame, fpath,
                              self._remove_file, even=(i % 2 == 0))
                row.pack(fill='x', pady=1)
        self._list_frame.update_idletasks()

    def _update_count(self):
        n = len(self._files)
        self.count_var.set(f'{n} file{"s" if n != 1 else ""} queued')

    def _on_list_resize(self, e):
        self._canvas.configure(scrollregion=self._canvas.bbox('all'))

    def _on_canvas_resize(self, e):
        self._canvas.itemconfig(self._canvas_window, width=e.width)

    def _on_scroll(self, e):
        delta = -1 * (e.delta // 120) if IS_WIN else (-1 if e.delta > 0 else 1)
        self._canvas.yview_scroll(delta, 'units')

    # ── Output path ────────────────────────────────────────────────────────

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title='Save Excel file as…',
            defaultextension='.xlsx',
            initialfile='fs2_messages.xlsx',
            filetypes=[('Excel Workbook', '*.xlsx'), ('All files', '*.*')])
        if path:
            self.out_path.set(path)

    # ── Run ────────────────────────────────────────────────────────────────

    def _run(self):
        if not HAVE_OPENPYXL:
            messagebox.showerror('Missing dependency',
                                 'openpyxl is required.\nRun: pip install openpyxl')
            return
        if not self._files:
            messagebox.showwarning('No files', 'Add at least one .fs2 file first.')
            return
        out = self.out_path.get().strip()
        if not out:
            messagebox.showwarning('No output path', 'Choose an output .xlsx path first.')
            return

        self._running = True
        self.run_btn.config(state='disabled')
        self.progress['value'] = 0
        self._result_rows.clear()
        self.result_var.set('')
        self._log('─' * 50, 'dim')
        self._log(f'Starting export of {len(self._files)} file(s)…', 'hdr')

        def worker():
            all_rows = []
            errors = []
            total = len(self._files)
            for i, fpath in enumerate(self._files):
                pct = int((i / total) * 80)
                self.after(0, lambda p=pct: self.progress.config(value=p))
                self.after(0, lambda f=fpath:
                           self.status_var.set(f'Processing {Path(f).name}…'))
                mnum, mname, rows, err = extract_file(fpath)
                if err:
                    errors.append(f'{Path(fpath).name}: {err}')
                    self.after(0, lambda e=err, f=fpath:
                               self._log(f'  ✗  {Path(f).name}  —  {e}', 'err'))
                else:
                    label = f'M{mnum}: {mname}' if mnum else Path(fpath).stem
                    self.after(0, lambda l=label, n=len(rows):
                               self._log(f'  ✓  {l:<42} {n:>4} messages', 'ok'))
                    all_rows.extend(rows)

            self.after(0, lambda: self.progress.config(value=85))
            self.after(0, lambda: self.status_var.set('Building spreadsheet…'))
            self.after(0, lambda: self._log('Building workbook…', 'dim'))

            try:
                wb = build_workbook(all_rows)
                wb.save(out)
                self.after(0, lambda: self._on_done(all_rows, errors, out))
            except Exception as ex:
                self.after(0, lambda e=str(ex): self._on_error(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_done(self, rows, errors, out):
        self._running = False
        self._result_rows = rows
        self.run_btn.config(state='normal')
        self.progress['value'] = 100
        self.status_var.set(f'Done — {len(rows)} rows exported.')

        with_sender = sum(1 for r in rows if r['sender'])
        self._log('─' * 50, 'dim')
        self._log(f'Total rows:         {len(rows)}', 'hdr')
        self._log(f'With sender:        {with_sender}', 'ok')
        self._log(f'Without sender:     {len(rows) - with_sender}', 'warn')
        if errors:
            self._log(f'Errors:             {len(errors)}', 'err')
        self._log(f'Saved → {out}', 'ok')

        self.result_var.set(
            f'{len(rows)} rows  ·  {with_sender} with sender  ·  '
            f'{len(rows)-with_sender} unconfirmed  ·  {len(errors)} errors')

        if messagebox.askyesno('Export complete',
                               f'Exported {len(rows)} rows to:\n{out}\n\n'
                               f'Open the file now?'):
            self._open_file(out)

    def _on_error(self, msg):
        self._running = False
        self.run_btn.config(state='normal')
        self.progress['value'] = 0
        self._log(f'FATAL ERROR: {msg}', 'err')
        messagebox.showerror('Export failed', f'An error occurred:\n\n{msg}')

    def _open_file(self, path):
        import subprocess
        try:
            if IS_WIN:
                os.startfile(path)
            elif IS_MAC:
                subprocess.call(['open', path])
            else:
                subprocess.call(['xdg-open', path])
        except Exception:
            pass

    # ── Log helpers ────────────────────────────────────────────────────────

    def _log(self, text, tag=None):
        self.log.config(state='normal')
        self.log.insert('end', text + '\n', tag or '')
        self.log.see('end')
        self.log.config(state='disabled')

    def _save_log(self):
        path = filedialog.asksaveasfilename(
            defaultextension='.txt',
            initialfile='fs2_export_log.txt',
            filetypes=[('Text file', '*.txt'), ('All files', '*.*')])
        if path:
            content = self.log.get('1.0', 'end')
            Path(path).write_text(content, encoding='utf-8')
            self._log(f'Log saved → {path}', 'dim')


# ═════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    app = App()
    app.mainloop()
